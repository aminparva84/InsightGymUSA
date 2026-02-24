"""
Export ACTION_JSON_SCENARIOS to Excel (.xlsx) and CSV (UTF-8 with BOM).
Run: python scripts/export_action_json_scenarios.py
Output: ACTION_JSON_SCENARIOS.xlsx and ACTION_JSON_SCENARIOS.csv in project root.
CSV uses UTF-8 BOM so Excel and other tools display Persian correctly.
"""

import csv
import os
import sys

# Add backend to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

SCENARIOS = [
    # Role, Scenario, Action JSON, Parameters, Example FA, Example EN, Status
    ('Member', 'Search exercises', 'search_exercises', 'query?, target_muscle?, level?, intensity?, max_results?, language?', 'تمرینات سینه', 'chest exercises', 'Ready'),
    ('Member', 'Search exercises by target muscle', 'search_exercises', 'target_muscle, query?, max_results?', 'تمرینات برای عضله پا', 'exercises for legs', 'Ready'),
    ('Member', 'Buy or get training plan suggestions', 'suggest_training_plans', 'language?, max_results?', 'میخوام برنامه تمرینی بخرم / چی پیشنهاد میدی؟', 'I want to buy a training plan / what do you suggest?', 'Ready'),
    ('Member', 'Generate workout after purchase', 'create_workout_plan', 'month?, target_muscle?, language?', 'برنامه‌ام رو بساز / برنامه خریدم بساز', 'generate my workout / build my program', 'Ready'),
    ('Member', 'Update profile (age weight goals etc)', 'update_user_profile', 'fields: {age?, weight?, height?, fitness_goals?, injuries?, ...}', 'سن من ۲۵ است / هدفم افزایش عضله', 'my age is 25 / my goal is muscle gain', 'Ready'),
    ('Member', 'Request progress check from trainer', 'progress_check', "mode: 'request'", 'درخواست بررسی پیشرفت', 'request progress check', 'Ready'),
    ('Member', 'Send message to trainer', 'trainer_message', 'body, recipient_id? (auto if member)', 'پیام به مربی', 'message to trainer', 'Ready'),
    ('Member', 'Ask about BMI / weight / progress', 'get_dashboard_progress', 'language?, fields?', 'BMI من چنده؟ / وزنم چقدره؟ / پیشرفتم', 'what is my BMI? / my weight? / my progress', 'Ready'),
    ('Member', 'Add weight to Progress Trend', 'add_progress_entry', 'weight_kg?, chest_cm?, waist_cm?, hips_cm?, arm_left_cm?, arm_right_cm?, thigh_left_cm?, thigh_right_cm?', 'وزن ۷۶ کیلو اضافه کن', 'add my weight 76 kg', 'Ready'),
    ('Member', 'Ask what is my training today', 'get_todays_training', 'language?', 'جلسه امروز چیه؟ / برنامه امروز / تمرین امروز', 'what is my training today? / today workout', 'Ready'),
    ('Member', 'Ask about Psychology Test', 'get_dashboard_tab_info', "tab: 'psychology-test', language?", 'تست روانشناسی چیه؟', 'what is Psychology Test?', 'Ready'),
    ('Member', 'Ask about Online Laboratory', 'get_dashboard_tab_info', "tab: 'online-lab', language?", 'آزمایشگاه آنلاین چیه؟ / ماشین حساب BMI', 'what is Online Laboratory? / BMI calculator', 'Ready'),
    ('Member', 'Schedule meeting', 'schedule_meeting', 'appointment_date?, appointment_time?, duration?, notes?', 'جلسه فردا صبح', 'schedule meeting tomorrow morning', 'Ready'),
    ('Member', 'Schedule appointment', 'schedule_appointment', 'appointment_date?, appointment_time?, duration?, notes?', 'نوبت فردا عصر', 'appointment tomorrow afternoon', 'Ready'),
    ('Assistant', 'Respond to progress check', 'progress_check', "mode: 'respond', request_id, status: 'accepted'|'denied'", 'تایید درخواست پیشرفت', 'accept progress check request', 'Ready'),
    ('Assistant', 'Send message to assigned member', 'trainer_message', 'recipient_id, body', 'پیام به عضو', 'message to member', 'Ready'),
    ('Assistant', 'Schedule meeting with member', 'schedule_meeting', 'appointment_date?, appointment_time?, duration?, notes?', 'جلسه با عضو', 'schedule with member', 'Ready'),
    ('Assistant', 'Ask about trainers / assistants list', 'get_trainers_info', 'language?', 'لیست مربیان / دستیاران / اعضای من', 'list of trainers / my assigned members', 'Ready'),
    ('Assistant', "Ask about a member's progress", 'get_member_progress', 'member_id?, member_username?, language?', 'وضعیت عضو علی / پیشرفت عضو X', "check member ali's progress / member X situation", 'Ready'),
    ('Admin', 'Update site settings', 'site_settings', 'fields: {contact_email?, contact_phone?, address_fa?, address_en?, ...}', 'تنظیمات سایت', 'update site settings', 'Ready'),
    ('Admin', 'Ask about trainers / assistants list', 'get_trainers_info', 'language?', 'لیست مربیان / دستیاران', 'list of trainers / assistants', 'Ready'),
    ('Admin', "Ask about any member's progress", 'get_member_progress', 'member_id?, member_username?, language?', 'وضعیت عضو X / پیشرفت عضو', 'member progress / check member situation', 'Ready'),
    ('Admin', 'All assistant actions', '(same as Assistant)', '(same as Assistant)', '-', '-', 'Ready'),
    ('Admin', 'Search exercises', 'search_exercises', 'query?, target_muscle?, ...', 'جستجوی تمرینات', 'search exercises', 'Ready'),
    ('Admin', 'Suggest training plans', 'suggest_training_plans', 'language?, max_results?', 'پیشنهاد برنامه', 'suggest plans', 'Ready'),
    ('Guest', 'Open AI chat', 'respond', '(chat UI - login required)', '-', '-', 'Ready'),
    ('Guest', 'Register', 'respond', '(UI modal - no action)', '-', '-', 'Ready'),
    ('Guest', 'Login', 'respond', '(UI modal - no action)', '-', '-', 'Ready'),
    ('All', 'Navigate to dashboard', 'navigate', 'path: /dashboard', 'برو به داشبورد', 'go to dashboard', 'Ready'),
    ('All', 'Navigate to training program', 'navigate', 'path: /dashboard?tab=training-program', 'برنامه تمرینی', 'training program', 'Ready'),
    ('All', 'Navigate to profile', 'navigate', 'path: /dashboard?tab=profile', 'پروفایل', 'profile', 'Ready'),
    ('All', 'Navigate to progress trend', 'navigate', 'path: /dashboard?tab=progress-trend', 'روند تغییرات', 'progress trend', 'Ready'),
    ('All', 'Navigate to psychology test', 'navigate', 'path: /dashboard?tab=psychology-test', 'تست روانشناسی', 'psychology test', 'Ready'),
    ('All', 'Navigate to online lab', 'navigate', 'path: /dashboard?tab=online-lab', 'آزمایشگاه آنلاین', 'online laboratory', 'Ready'),
]


def build_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('Install: pip install openpyxl')
        raise

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Action JSON Scenarios'

    headers = ['Role', 'Scenario', 'Action JSON', 'Parameters', 'Example User Message (FA)', 'Example User Message (EN)', 'Status']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, row_data in enumerate(SCENARIOS, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 8

    out_path = os.path.join(os.path.dirname(__file__), '..', 'ACTION_JSON_SCENARIOS.xlsx')
    wb.save(out_path)
    print(f'Saved: {out_path}')
    return out_path


def build_csv():
    """Export scenarios to CSV with UTF-8 BOM for correct Persian display in Excel."""
    headers = ['Role', 'Scenario', 'Action JSON', 'Parameters', 'Example User Message (FA)', 'Example User Message (EN)', 'Status']
    out_path = os.path.join(os.path.dirname(__file__), '..', 'ACTION_JSON_SCENARIOS.csv')
    with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(SCENARIOS)
    print(f'Saved: {out_path} (UTF-8 with BOM)')
    return out_path


if __name__ == '__main__':
    build_excel()
    build_csv()
