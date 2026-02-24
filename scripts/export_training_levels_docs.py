"""
Export Training Level Info tab content to Excel and PDF for admin review/confirmation.
Run from project root: python scripts/export_training_levels_docs.py
Output: docs/Training_Level_Info_Export.xlsx and docs/Training_Level_Info_Export.pdf
"""
import os
import sys

# Ensure project root is on path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'docs')
EXCEL_PATH = os.path.join(OUTPUT_DIR, 'Training_Level_Info_Export.xlsx')
PDF_PATH = os.path.join(OUTPUT_DIR, 'Training_Level_Info_Export.pdf')

# --- Training Level Info (from admin confirmation) ---

TRAINING_LEVELS_DATA = {
    'beginner': {
        'label_en': 'Beginner',
        'label_fa': 'مبتدی',
        'description_en': 'Beginner user: 0–6 months training history; low control over movement form; little familiarity with breathing and mind–muscle coordination; fast recovery but injury-sensitive. Main goals: learn correct form, activate muscles, increase body awareness, build training habit, prevent injury. Design: choose simple, safe, controllable exercises; focus on form, breathing, correct range of motion. General program: 2–4 sessions/week; simple compound movements, machine or bodyweight; 2–3 sets, 10–15 reps; 60–90 sec rest; no advanced techniques (superset, drop set). Breathing & mindfulness: breathing cue per movement; inhale on negative phase, exhale on positive; short pause for body awareness.',
        'description_fa': 'تعریف کاربر مبتدی: سابقه تمرین ۰ تا ۶ ماه؛ کنترل پایین روی فرم حرکات؛ آشنایی کم با تنفس و هماهنگی ذهن-عضله؛ ریکاوری سریع اما حساس به آسیب. اهداف اصلی: یادگیری فرم صحیح حرکات، فعال‌سازی عضلات، افزایش آگاهی بدنی، ایجاد عادت تمرینی، جلوگیری از آسیب. طراحی تمرین: تمرین‌ها را ساده، ایمن و قابل کنترل انتخاب کند؛ تمرکز روی فرم حرکت، تنفس، دامنه حرکتی صحیح. ویژگی کلی: ۲ تا ۴ جلسه در هفته؛ حرکات چندمفصلی ساده، دستگاه یا وزن بدن؛ ۲–۳ ست، ۱۰–۱۵ تکرار؛ استراحت ۶۰–۹۰ ثانیه؛ بدون تکنیک‌های پیشرفته (سوپرست، دراپ‌ست و…). تنفس و مایندفولنس: راهنمای تنفس در هر حرکت؛ تاکید روی دم در فاز منفی و بازدم در فاز مثبت؛ مکث کوتاه برای آگاهی بدن.',
        'goals': [
            {'en': 'Learn correct movement form', 'fa': 'یادگیری فرم صحیح حرکات'},
            {'en': 'Activate muscles', 'fa': 'فعال‌سازی عضلات'},
            {'en': 'Increase body awareness', 'fa': 'افزایش آگاهی بدنی'},
            {'en': 'Build training habit', 'fa': 'ایجاد عادت تمرینی'},
            {'en': 'Prevent injury', 'fa': 'جلوگیری از آسیب'},
        ],
        'purposes': {
            'gain_weight': {
                'sessions_per_week': '3',
                'sets_per_action': '2–3',
                'reps_per_action': '12–15',
                'training_focus_en': 'Full-body; light to moderate intensity. Weight gain via muscle activation + appetite, not high load.',
                'training_focus_fa': 'فول‌بادی؛ شدت سبک تا متوسط. افزایش وزن از طریق فعال‌سازی عضلانی + اشتها نه فشار بالا.',
                'break_between_sets': '60–90 sec',
            },
            'lose_weight': {
                'sessions_per_week': '3–4',
                'sets_per_action': '2–3',
                'reps_per_action': '15–20',
                'training_focus_en': 'Full-body + light metabolic; fat loss without nervous fatigue, with calm and breathing.',
                'training_focus_fa': 'فول‌بادی + متابولیک سبک؛ چربی‌سوزی بدون خستگی عصبی، با حفظ آرامش و تنفس.',
                'break_between_sets': '30–60 sec',
            },
            'gain_muscle': {
                'sessions_per_week': '3',
                'sets_per_action': '3',
                'reps_per_action': '10–12',
                'training_focus_en': 'Full-body; focus on mind–muscle connection, not load volume.',
                'training_focus_fa': 'فول‌بادی؛ تمرکز روی اتصال ذهن–عضله نه حجم وزنه.',
                'break_between_sets': '60–90 sec',
            },
            'shape_fitting': {
                'sessions_per_week': '3–4',
                'sets_per_action': '2–3',
                'reps_per_action': '12–15',
                'training_focus_en': 'Combined; smooth, continuous movements, focus on form.',
                'training_focus_fa': 'ترکیبی؛ حرکات نرم، پیوسته، با تمرکز روی فرم.',
                'break_between_sets': '45–60 sec',
            },
        },
    },
    'intermediate': {
        'label_en': 'Intermediate',
        'label_fa': 'متوسط',
        'description_en': 'Intermediate user: 6 months–2 years training; relative familiarity with form; ability to maintain focus; tolerance for volume; ready for variety. Main goals: increase relative strength, muscle shaping, improve nerve–muscle coordination, introduce training variety. Design: increase intensity, add movement variety, take body out of adaptation. General program: 3–5 sessions/week; bodybuilding + functional; 3–4 sets, 8–12 reps; 45–75 sec rest; limited use of superset, compound sets. Breathing & focus: mind–muscle connection, movement rhythm control, breathing in sync with effort.',
        'description_fa': 'تعریف کاربر متوسط: سابقه تمرین ۶ ماه تا ۲ سال؛ آشنایی نسبی با فرم؛ توانایی حفظ تمرکز در تمرین؛ توان تحمل حجم تمرین؛ آمادگی برای تنوع. اهداف اصلی: افزایش قدرت نسبی، فرم‌دهی عضلانی، افزایش هماهنگی عصب-عضله، شروع تنوع تمرینی. منطق طراحی: شدت تمرین را افزایش دهد؛ تنوع حرکتی ایجاد کند؛ بدن را از حالت سازگاری خارج کند. ویژگی‌های کلی: ۳ تا ۵ جلسه در هفته؛ ترکیب بدنسازی + فانکشنال؛ ۳–۴ ست، ۸–۱۲ تکرار؛ استراحت ۴۵–۷۵ ثانیه؛ استفاده محدود از سوپرست، تمرینات ترکیبی. تنفس و تمرکز: تمرکز روی اتصال ذهن-عضله؛ کنترل ریتم حرکت؛ هماهنگی تنفس با فشار تمرین.',
        'goals': [
            {'en': 'Increase relative strength', 'fa': 'افزایش قدرت نسبی'},
            {'en': 'Muscle shaping', 'fa': 'فرم‌دهی عضلانی'},
            {'en': 'Improve nerve–muscle coordination', 'fa': 'افزایش هماهنگی عصب-عضله'},
            {'en': 'Introduce training variety', 'fa': 'شروع تنوع تمرینی'},
        ],
        'purposes': {
            'gain_weight': {
                'sessions_per_week': '4–5',
                'sets_per_action': '3–4',
                'reps_per_action': '6–10',
                'training_focus_en': 'Split by muscle group; bodybuilding + functional; 45–75 sec rest.',
                'training_focus_fa': 'اسپلیت گروه عضلانی؛ بدنسازی + فانکشنال؛ استراحت ۴۵–۷۵ ثانیه.',
                'break_between_sets': '45–75 sec',
            },
            'lose_weight': {
                'sessions_per_week': '4–5',
                'sets_per_action': '3–4',
                'reps_per_action': '10–12',
                'training_focus_en': 'Upper/lower or metabolic + strength; limited superset.',
                'training_focus_fa': 'بالا/پایین یا متابولیک + قدرتی؛ سوپرست محدود.',
                'break_between_sets': '45–60 sec',
            },
            'gain_muscle': {
                'sessions_per_week': '4–5',
                'sets_per_action': '3–4',
                'reps_per_action': '8–12',
                'training_focus_en': 'Bodybuilding + functional; mind–muscle connection.',
                'training_focus_fa': 'بدنسازی + فانکشنال؛ اتصال ذهن-عضله.',
                'break_between_sets': '60–75 sec',
            },
            'shape_fitting': {
                'sessions_per_week': '3–5',
                'sets_per_action': '3–4',
                'reps_per_action': '10–15',
                'training_focus_en': 'Combined; movement quality, breathing control.',
                'training_focus_fa': 'ترکیبی؛ کیفیت حرکت، کنترل تنفس.',
                'break_between_sets': '45–60 sec',
            },
        },
    },
    'advanced': {
        'label_en': 'Advanced',
        'label_fa': 'پیشرفته',
        'description_en': 'Advanced user: over 2 years training; high mastery of form; high physical and mental readiness; high volume tolerance; breathing control; strong mind–muscle connection. Main goals: targeted hypertrophy or strength, advanced fat loss, neuromuscular challenge, performance improvement. Design: fully personalize training; use advanced techniques; intelligently manage intensity, volume, recovery. Program: 4–6 sessions/week; professional mix of bodybuilding, functional, neural training; variable sets/reps (6–15); rest 30–60 sec or variable; techniques: superset, drop set, time under tension, metabolic training. Breathing & mindfulness: breathing control under load; deep focus on body energy; mindful training.',
        'description_fa': 'تعریف کاربر پیشرفته: سابقه تمرین بیش از ۲ سال؛ تسلط بالا روی فرم؛ آمادگی بدنی و ذهنی بالا؛ آمادگی حجم بالا؛ کنترل تنفس؛ اتصال ذهن و عضله قوی. اهداف اصلی: افزایش حجم یا قدرت هدفمند؛ چربی‌سوزی پیشرفته؛ چالش عصبی-عضلانی؛ بهبود عملکرد. منطق طراحی: تمرین‌ها را کاملاً شخصی‌سازی کند؛ از تکنیک‌های پیشرفته استفاده کند؛ شدت، حجم و ریکاوری را هوشمندانه تنظیم کند. ویژگی‌های برنامه: ۴ تا ۶ جلسه در هفته؛ ترکیب حرفه‌ای بدنسازی، فانکشنال، تمرینات عصبی؛ ست و تکرار متغیر (۶–۱۵)؛ استراحت ۳۰–۶۰ ثانیه یا متغیر؛ تکنیک‌ها: سوپرست، دراپ‌ست، تایم آندر تنشن، تمرینات متابولیک. تنفس و مایندفولنس: کنترل تنفس تحت فشار؛ تمرکز عمیق روی انرژی بدن؛ تمرین آگاهانه و حضور ذهن.',
        'goals': [
            {'en': 'Targeted hypertrophy or strength', 'fa': 'افزایش حجم یا قدرت هدفمند'},
            {'en': 'Advanced fat loss', 'fa': 'چربی‌سوزی پیشرفته'},
            {'en': 'Neuromuscular challenge', 'fa': 'چالش عصبی-عضلانی'},
            {'en': 'Performance improvement', 'fa': 'بهبود عملکرد'},
        ],
        'purposes': {
            'gain_weight': {
                'sessions_per_week': '5',
                'sets_per_action': '4–5',
                'reps_per_action': '6–10',
                'training_focus_en': 'Advanced split; 90–120 sec rest.',
                'training_focus_fa': 'اسپلیت پیشرفته؛ استراحت ۹۰–۱۲۰ ثانیه.',
                'break_between_sets': '90–120 sec',
            },
            'lose_weight': {
                'sessions_per_week': '5–6',
                'sets_per_action': '3–4',
                'reps_per_action': '10–15',
                'training_focus_en': 'Metabolic + strength; superset / giant set; 30–45 sec rest.',
                'training_focus_fa': 'تمرین متابولیک + قدرتی؛ سوپرست / جاینت‌ست؛ استراحت ۳۰–۴۵ ثانیه.',
                'break_between_sets': '30–45 sec',
            },
            'gain_muscle': {
                'sessions_per_week': '5',
                'sets_per_action': '4–5',
                'reps_per_action': '6–12',
                'training_focus_en': 'Volume + control; TUT, limited drop set; 60–90 sec rest.',
                'training_focus_fa': 'حجم + کنترل؛ TUT، دراپ‌ست محدود؛ استراحت ۶۰–۹۰ ثانیه.',
                'break_between_sets': '60–90 sec',
            },
            'shape_fitting': {
                'sessions_per_week': '4–5',
                'sets_per_action': '3–4',
                'reps_per_action': '12–15',
                'training_focus_en': 'Movement quality; 30–60 sec rest.',
                'training_focus_fa': 'کیفیت حرکت؛ استراحت ۳۰–۶۰ ثانیه.',
                'break_between_sets': '30–60 sec',
            },
        },
    },
}

PURPOSE_LABELS = {
    'lose_weight': ('Lose weight', 'کاهش وزن'),
    'gain_weight': ('Gain weight', 'افزایش وزن'),
    'gain_muscle': ('Gain muscle', 'افزایش عضله'),
    'shape_fitting': ('Shape fitting', 'تناسب اندام'),
}

# Labels for purpose fields (EN, FA) – used in Excel/PDF
PURPOSE_FIELD_LABELS = {
    'sessions_per_week': ('Sessions per week', 'تعداد جلسات در هفته'),
    'sets_per_action': ('Sets per action', 'تعداد ست در هر حرکت'),
    'reps_per_action': ('Reps per action', 'تعداد تکرار در هر حرکت'),
    'training_focus': ('Training focus (where & tools)', 'نحوه تمرین (مکان و ابزار)'),
    'break_between_sets': ('Break between sets', 'استراحت بین ست‌ها'),
}
SECTION_LABELS = {
    'training_level_info': ('Training Level Information', 'اطلاعات سطح‌های تمرینی'),
    'corrective_movements': ('Corrective movements for each injury', 'حرکات اصلاحی برای هر آسیب'),
    'sub_section': ('Sub-section', 'زیربخش'),
    'content_en': ('Content (EN)', 'محتوا (انگلیسی)'),
    'content_fa': ('Content (FA)', 'محتوا (فارسی)'),
    'description': ('Description', 'توضیحات'),
    'goals': ('Goals', 'اهداف'),
    'features_per_purpose': ('Features for each training purpose', 'ویژگی‌ها برای هر هدف تمرینی'),
    'purposes': ('Purposes / Description', 'اهداف / کاربرد'),
    'allowed_movements': ('Allowed movements', 'حرکات مجاز'),
    'forbidden_movements': ('Forbidden movements', 'حرکات ممنوع'),
    'important_notes': ('Important notes', 'نکات مهم'),
    'common_note': ('Common note for all injuries', 'نکته مشترک برای تمام آسیب‌ها'),
}

# Corrective movements: main goals, allowed, forbidden, important notes (from user confirmation)
COMMON_INJURY_NOTE = {
    'en': 'Common note for all injuries: Always correct form + controlled breathing. If pain is severe → stop the movement and substitute. Safety first, then performance. Remove direct load on the injured joint. Substitute the movement, not remove the whole workout. Focus on: safe range of motion, control, breathing.',
    'fa': 'نکته مشترک برای تمام آسیب‌ها: همیشه فرم صحیح + تنفس کنترل‌شده. اگر درد شدید بود → توقف حرکت و جایگزینی. اول ایمنی، بعد عملکرد. حذف فشار مستقیم روی مفصل آسیب‌دیده. جایگزینی حرکت، نه حذف کل تمرین. تمرکز روی: دامنه امن، کنترل، تنفس.',
}

INJURIES_DATA = {
    'knee': {
        'label_en': 'Knee',
        'label_fa': 'زانو',
        'purposes_en': 'Main goals during knee injury: 1. Reduce direct pressure on the knee joint. 2. Prevent further injury or worsening pain. 3. Strengthen supporting muscles (quadriceps, hamstrings, glutes). 4. Maintain safe range of motion. 5. Maintain body control and correct movement form.',
        'purposes_fa': 'اهداف اصلی در هنگام آسیب زانو: 1. کاهش فشار مستقیم روی مفصل زانو. 2. جلوگیری از آسیب بیشتر یا تشدید درد. 3. تقویت عضلات حمایت‌کننده (کوادری‌سپس، همسترینگ، گلوت‌ها). 4. حفظ دامنه حرکتی امن. 5. حفظ کنترل بدن و فرم درست حرکت.',
        'allowed_movements': [
            {'en': 'Glute Bridge → strengthen glutes and hamstrings without knee load', 'fa': 'Glute Bridge → تقویت باسن و همسترینگ بدون فشار روی زانو'},
            {'en': 'Hip Thrust → same as Glute Bridge, load on glutes', 'fa': 'Hip Thrust → همانند Glute Bridge، بار روی باسن'},
            {'en': 'Hamstring Curl (machine or ball) → strengthen posterior thigh without front knee pressure', 'fa': 'Hamstring Curl (دستگاه یا توپ) → تقویت پشت ران بدون فشار جلو زانو'},
            {'en': 'Step-back Lunge or short Reverse Lunge → limited range', 'fa': 'Step-back Lunge یا Reverse Lunge کوتاه → لنج به عقب با دامنه محدود'},
            {'en': 'Short Wall Sit → stability without heavy load', 'fa': 'Wall Sit کوتاه → حفظ ثبات بدون فشار زیاد'},
            {'en': 'Light Leg Press with limited ROM → if machine is adjustable', 'fa': 'Leg Press سبک با دامنه محدود → اگر دستگاه قابل تنظیم باشد'},
            {'en': 'Isometric quad exercises → e.g. Quad Set', 'fa': 'تمرینات ایزومتریک برای کوادری‌سپس → مثل Quad Set'},
        ],
        'forbidden_movements': [
            {'en': 'Deep or fast Squat (Deep Squat)', 'fa': 'Squat عمیق یا سریع (Deep Squat)'},
            {'en': 'Jump Squat / Plyometric / severe jumping', 'fa': 'Jump Squat / Plyometric / حرکات پرشی شدید'},
            {'en': 'Long forward Lunge', 'fa': 'Lunge جلو بلند'},
            {'en': 'Heavy Leg Extension', 'fa': 'Leg Extension سنگین'},
            {'en': 'Rapid or sudden direction change', 'fa': 'تغییر جهت سریع و ناگهانی'},
            {'en': 'Explosive and high-speed movements', 'fa': 'حرکات انفجاری و سرعت بالا'},
        ],
        'important_notes_en': '1. If pain is moderate or severe → light leg work or substitute. 2. If user has multiple injuries → knee limitation takes priority. 3. Focus on movement control and breathing (calm inhale, exhale on effort phase). 4. Recovery time between sets should be slightly longer than usual. 5. In workout text state: "If pain worsens, stop the exercise and choose a substitute movement."',
        'important_notes_fa': '1. اگر درد متوسط یا شدید باشد → حرکات پا سبک یا جایگزین شوند. 2. اگر کاربر چند آسیب همزمان دارد → محدودیت زانو اولویت می‌گیرد. 3. تمرکز روی کنترل حرکت و تنفس (دم آرام، بازدم در فاز فشار). 4. زمان ریکاوری بین ست‌ها باید کمی بیشتر از حالت معمول باشد. 5. در متن تمرینی ذکر شود: «اگر درد تشدید شد، تمرین متوقف شود و حرکت جایگزین انتخاب گردد.»',
    },
    'shoulder': {
        'label_en': 'Shoulder',
        'label_fa': 'شانه',
        'purposes_en': 'Main goals in shoulder injury: 1. Maintain scapular stability. 2. Avoid pressure on rotator cuff. 3. Strengthen shoulder stabilizers.',
        'purposes_fa': 'اهداف اصلی در آسیب شانه: 1. حفظ ثبات کتف. 2. جلوگیری از فشار روی روتاتور کاف. 3. تقویت عضلات نگهدارنده شانه.',
        'allowed_movements': [
            {'en': 'Seated Row', 'fa': 'Seated Row'},
            {'en': 'Lat Pulldown (front)', 'fa': 'Lat Pulldown جلو'},
            {'en': 'Face Pull', 'fa': 'Face Pull'},
            {'en': 'Light Lateral Raise', 'fa': 'Lateral Raise سبک'},
            {'en': 'External Rotation', 'fa': 'External Rotation'},
            {'en': 'Scapular control exercises', 'fa': 'تمرینات کنترل کتف'},
        ],
        'forbidden_movements': [
            {'en': 'Heavy Overhead Press', 'fa': 'Overhead Press سنگین'},
            {'en': 'Upright Row', 'fa': 'Upright Row'},
            {'en': 'Dips', 'fa': 'Dips'},
            {'en': 'Behind-neck Pull-up', 'fa': 'Pull-up پشت گردن'},
            {'en': 'Heavy Bench Press', 'fa': 'Bench Press سنگین'},
            {'en': 'Explosive overhead movements', 'fa': 'حرکات انفجاری بالای سر'},
        ],
        'important_notes_en': 'Limited range of motion. Focus on scapular stability. Exhale on effort, inhale on return.',
        'important_notes_fa': 'دامنه حرکتی محدود. تمرکز روی ثبات کتف. بازدم هنگام فشار، دم هنگام برگشت.',
    },
    'lower_back': {
        'label_en': 'Lower back',
        'label_fa': 'کمر',
        'purposes_en': 'Main goals during lower back injury: 1. Reduce axial load. 2. Increase core stability. 3. Prevent worsening pain.',
        'purposes_fa': 'اهداف اصلی در هنگام آسیب کمر: 1. کاهش فشار محوری. 2. افزایش ثبات مرکزی (Core Stability). 3. جلوگیری از تشدید درد.',
        'allowed_movements': [
            {'en': 'Bird Dog', 'fa': 'Bird Dog'},
            {'en': 'Dead Bug', 'fa': 'Dead Bug'},
            {'en': 'Glute Bridge', 'fa': 'Glute Bridge'},
            {'en': 'Light Hip Hinge', 'fa': 'Hip Hinge سبک'},
            {'en': 'Modified Plank', 'fa': 'Plank اصلاح‌شده'},
            {'en': 'Light Cable Pull-through', 'fa': 'Cable Pull-through سبک'},
        ],
        'forbidden_movements': [
            {'en': 'Heavy Deadlift', 'fa': 'Deadlift سنگین'},
            {'en': 'Good Morning', 'fa': 'Good Morning'},
            {'en': 'Back Squat', 'fa': 'Back Squat'},
            {'en': 'Russian Twist', 'fa': 'Russian Twist'},
            {'en': 'Hyperextension', 'fa': 'Hyperextension'},
            {'en': 'Sudden rotational movements', 'fa': 'حرکات چرخشی ناگهانی'},
        ],
        'important_notes_en': 'Controlled breathing. Slow movement, precise form. Focus on core muscles.',
        'important_notes_fa': 'تنفس کنترل‌شده. حرکت آرام، فرم دقیق. تمرکز روی عضلات مرکزی.',
    },
    'neck': {
        'label_en': 'Neck',
        'label_fa': 'گردن',
        'purposes_en': 'Main goals: 1. Remove tension. 2. Maintain neutral spine. 3. Avoid direct pressure.',
        'purposes_fa': 'اهداف اصلی: 1. حذف تنش. 2. حفظ حالت خنثی ستون فقرات. 3. جلوگیری از فشار مستقیم.',
        'allowed_movements': [
            {'en': 'Chin Tuck', 'fa': 'Chin Tuck'},
            {'en': 'Isometric neck exercises', 'fa': 'تمرینات ایزومتریک گردن'},
            {'en': 'Upper-body work with neutral neck', 'fa': 'تمرینات بالاتنه با گردن خنثی'},
            {'en': 'Light breathing exercises', 'fa': 'تمرینات سبک تنفسی'},
        ],
        'forbidden_movements': [
            {'en': 'Heavy Shrug', 'fa': 'Shrug سنگین'},
            {'en': 'Explosive movements', 'fa': 'حرکات انفجاری'},
            {'en': 'Direct pressure on head', 'fa': 'فشار مستقیم روی سر'},
            {'en': 'Rapid neck rotation', 'fa': 'چرخش سریع گردن'},
        ],
        'important_notes_en': 'Full neck control. Breathing in sync with movement. If pain → substitute movement.',
        'important_notes_fa': 'کنترل کامل گردن. هماهنگی تنفس با حرکت. اگر درد باشد → جایگزین حرکات.',
    },
    'wrist': {
        'label_en': 'Wrist',
        'label_fa': 'مچ دست',
        'purposes_en': 'Main goals: 1. Reduce wrist flexion/extension load. 2. Maintain safe movement. 3. Strengthen forearm and prevent pain.',
        'purposes_fa': 'اهداف اصلی: 1. کاهش فشار خم و باز شدن مچ. 2. حفظ حرکت امن. 3. تقویت ساعد و پیشگیری از درد.',
        'allowed_movements': [
            {'en': 'Dumbbell Neutral Grip Press', 'fa': 'Dumbbell Neutral Grip Press'},
            {'en': 'Cable Push / Pull', 'fa': 'Cable Push / Pull'},
            {'en': 'Resistance Band Exercises', 'fa': 'Resistance Band Exercises'},
            {'en': 'Light forearm exercises', 'fa': 'تمرینات ساعد سبک'},
            {'en': 'Exercises without weight on hands', 'fa': 'تمرینات بدون تحمل وزن روی دست'},
        ],
        'forbidden_movements': [
            {'en': 'Classic Push-up', 'fa': 'Push-up کلاسیک'},
            {'en': 'Plank on hands', 'fa': 'Plank روی دست'},
            {'en': 'Barbell Press', 'fa': 'Barbell Press'},
            {'en': 'Heavy Hanging', 'fa': 'Hanging سنگین'},
            {'en': 'Burpee', 'fa': 'Burpee'},
        ],
        'important_notes_en': 'Limited range of motion. Use neutral grip. Control breathing under load.',
        'important_notes_fa': 'دامنه حرکت محدود. استفاده از گریپ خنثی. کنترل تنفس در فشار.',
    },
    'ankle': {
        'label_en': 'Ankle',
        'label_fa': 'مچ پا',
        'purposes_en': 'Main goals: 1. Joint stability and control. 2. Prevent sprain or direct pressure.',
        'purposes_fa': 'اهداف اصلی: 1. ثبات و کنترل مفصل. 2. جلوگیری از پیچ خوردگی یا فشار مستقیم.',
        'allowed_movements': [
            {'en': 'Seated Calf Raise', 'fa': 'Seated Calf Raise'},
            {'en': 'Glute-focused movements', 'fa': 'Glute-focused movements'},
            {'en': 'Short Step-up', 'fa': 'Step-up کوتاه'},
            {'en': 'Simple balance exercises', 'fa': 'تمرینات تعادلی ساده'},
            {'en': 'Controlled stretching', 'fa': 'تمرینات کششی کنترل‌شده'},
        ],
        'forbidden_movements': [
            {'en': 'Jumping / Plyometric', 'fa': 'Jumping / Plyometric'},
            {'en': 'High-intensity Running', 'fa': 'Running با شدت'},
            {'en': 'Box Jump', 'fa': 'Box Jump'},
            {'en': 'Rapid direction change', 'fa': 'تغییر جهت سریع'},
            {'en': 'Explosive single-leg exercises', 'fa': 'تمرینات تک‌پایی انفجاری'},
        ],
        'important_notes_en': 'Control range of motion. Focus on form and balance. Use calm inhale and exhale.',
        'important_notes_fa': 'کنترل دامنه حرکتی. تمرکز روی فرم و تعادل. استفاده از دم و بازدم آرام.',
    },
}


def build_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('Install: pip install openpyxl')
        raise

    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    section_font = Font(bold=True)

    # --- Sheet 1: Training Level Info ---
    sl, pf = SECTION_LABELS, PURPOSE_FIELD_LABELS
    ws1 = wb.active
    ws1.title = 'Training Level Info'
    row = 1
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws1.cell(row=row, column=1, value=f"{sl['training_level_info'][0]} / {sl['training_level_info'][1]}")
    cell.font = Font(bold=True, size=14)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    row += 2

    for level_key, level_data in TRAINING_LEVELS_DATA.items():
        ws1.cell(row=row, column=1, value=f"Level: {level_data['label_en']} / سطح: {level_data['label_fa']}")
        ws1.cell(row=row, column=1).font = section_font
        ws1.cell(row=row, column=1).fill = section_fill
        row += 1
        ws1.cell(row=row, column=1, value=f"{sl['sub_section'][0]} / {sl['sub_section'][1]}")
        ws1.cell(row=row, column=2, value=f"{sl['content_en'][0]} / {sl['content_en'][1]}")
        ws1.cell(row=row, column=3, value=f"{sl['content_fa'][0]} / {sl['content_fa'][1]}")
        for c in range(1, 4):
            ws1.cell(row=row, column=c).font = Font(bold=True)
        row += 1
        # Description
        ws1.cell(row=row, column=1, value=f"{sl['description'][0]} / {sl['description'][1]}")
        ws1.cell(row=row, column=2, value=level_data['description_en'])
        ws1.cell(row=row, column=3, value=level_data['description_fa'])
        row += 1
        # Goals
        ws1.cell(row=row, column=1, value=f"{sl['goals'][0]} / {sl['goals'][1]}")
        ws1.cell(row=row, column=2, value='; '.join(g['en'] for g in level_data['goals']))
        ws1.cell(row=row, column=3, value='; '.join(g['fa'] for g in level_data['goals']))
        row += 1
        # Features for each training purpose
        ws1.cell(row=row, column=1, value=f"{sl['features_per_purpose'][0]} / {sl['features_per_purpose'][1]}")
        ws1.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for purpose_key, purpose_label in PURPOSE_LABELS.items():
            p = level_data['purposes'][purpose_key]
            ws1.cell(row=row, column=1, value=f"  {purpose_label[0]} / {purpose_label[1]}")
            row += 1
            ws1.cell(row=row, column=1, value=f"    {pf['sessions_per_week'][0]} / {pf['sessions_per_week'][1]}")
            ws1.cell(row=row, column=2, value=p.get('sessions_per_week', ''))
            ws1.cell(row=row, column=3, value=p.get('sessions_per_week', ''))
            row += 1
            ws1.cell(row=row, column=1, value=f"    {pf['sets_per_action'][0]} / {pf['sets_per_action'][1]}")
            ws1.cell(row=row, column=2, value=p.get('sets_per_action', ''))
            ws1.cell(row=row, column=3, value=p.get('sets_per_action', ''))
            row += 1
            ws1.cell(row=row, column=1, value=f"    {pf['reps_per_action'][0]} / {pf['reps_per_action'][1]}")
            ws1.cell(row=row, column=2, value=p.get('reps_per_action', ''))
            ws1.cell(row=row, column=3, value=p.get('reps_per_action', ''))
            row += 1
            ws1.cell(row=row, column=1, value=f"    {pf['training_focus'][0]} / {pf['training_focus'][1]}")
            ws1.cell(row=row, column=2, value=p.get('training_focus_en', ''))
            ws1.cell(row=row, column=3, value=p.get('training_focus_fa', ''))
            row += 1
            ws1.cell(row=row, column=1, value=f"    {pf['break_between_sets'][0]} / {pf['break_between_sets'][1]}")
            ws1.cell(row=row, column=2, value=p.get('break_between_sets', ''))
            ws1.cell(row=row, column=3, value=p.get('break_between_sets', ''))
            row += 1
        row += 1

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 55
    ws1.column_dimensions['C'].width = 55

    # --- Sheet 2: Corrective movements (all labels EN + FA) ---
    ws2 = wb.create_sheet('Corrective Movements (Injuries)')
    row = 1
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws2.cell(row=row, column=1, value=f"{sl['corrective_movements'][0]} | {sl['corrective_movements'][1]}")
    cell.font = Font(bold=True, size=14)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    row += 2

    for injury_key, data in INJURIES_DATA.items():
        ws2.cell(row=row, column=1, value=f"Injury | آسیب: {data['label_en']} | {data['label_fa']}")
        ws2.cell(row=row, column=1).font = section_font
        ws2.cell(row=row, column=1).fill = section_fill
        row += 1
        ws2.cell(row=row, column=1, value=f"{sl['sub_section'][0]} | {sl['sub_section'][1]}")
        ws2.cell(row=row, column=2, value=f"{sl['content_en'][0]} | {sl['content_en'][1]}")
        ws2.cell(row=row, column=3, value=f"{sl['content_fa'][0]} | {sl['content_fa'][1]}")
        for c in range(1, 4):
            ws2.cell(row=row, column=c).font = Font(bold=True)
        row += 1
        ws2.cell(row=row, column=1, value=f"{sl['purposes'][0]} | {sl['purposes'][1]}")
        ws2.cell(row=row, column=2, value=data['purposes_en'])
        ws2.cell(row=row, column=3, value=data['purposes_fa'])
        row += 1
        ws2.cell(row=row, column=1, value=f"{sl['allowed_movements'][0]} | {sl['allowed_movements'][1]}")
        ws2.cell(row=row, column=2, value='; '.join(m['en'] for m in data['allowed_movements']))
        ws2.cell(row=row, column=3, value='; '.join(m['fa'] for m in data['allowed_movements']))
        row += 1
        ws2.cell(row=row, column=1, value=f"{sl['forbidden_movements'][0]} | {sl['forbidden_movements'][1]}")
        ws2.cell(row=row, column=2, value='; '.join(m['en'] for m in data['forbidden_movements']))
        ws2.cell(row=row, column=3, value='; '.join(m['fa'] for m in data['forbidden_movements']))
        row += 1
        ws2.cell(row=row, column=1, value=f"{sl['important_notes'][0]} | {sl['important_notes'][1]}")
        ws2.cell(row=row, column=2, value=data.get('important_notes_en', ''))
        ws2.cell(row=row, column=3, value=data.get('important_notes_fa', ''))
        row += 2

    ws2.cell(row=row, column=1, value=f"{sl['common_note'][0]} | {sl['common_note'][1]}")
    ws2.cell(row=row, column=1).font = section_font
    ws2.cell(row=row, column=1).fill = section_fill
    row += 1
    ws2.cell(row=row, column=2, value=COMMON_INJURY_NOTE['en'])
    ws2.cell(row=row, column=3, value=COMMON_INJURY_NOTE['fa'])
    row += 1

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 70
    ws2.column_dimensions['C'].width = 70

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(EXCEL_PATH)
    print('Excel saved:', EXCEL_PATH)


def _rtl_reshape(text):
    """Reshape Arabic/Persian for correct letter joining and return visual (RTL) order.
    Only applies to plain text; use _rtl_reshape_paragraph for strings that may contain HTML.
    """
    if not text or not isinstance(text, str):
        return text
    try:
        from arabic_reshaper import ArabicReshaper
        from bidi.algorithm import get_display
        reshaper = ArabicReshaper()  # default config: joining forms + ligatures
        reshaped = reshaper.reshape(text)
        return get_display(reshaped)
    except Exception:
        return text


def _rtl_reshape_paragraph(s):
    """Apply reshape+bidi only to text segments, not HTML tags, so ReportLab can parse HTML."""
    if not s or not isinstance(s, str):
        return s
    import re
    # Split by HTML tags but keep the tags
    parts = re.split(r'(<[^>]+>)', s)
    out = []
    for part in parts:
        if part.startswith('<') and part.endswith('>'):
            out.append(part)  # tag unchanged
        else:
            out.append(_rtl_reshape(part))  # reshape + bidi for text only
    return ''.join(out)


def _split_persian_description(text, max_chars=180):
    """Split long Persian/FA description into shorter chunks for cleaner RTL layout.
    Splits on sentence/clause boundaries (۔ ؛ .) then by length if still too long.
    """
    if not text or not isinstance(text, str):
        return [text] if text else []
    import re
    # Split on Persian period (۔), Arabic semicolon (؛), or period + space
    raw = re.split(r'(?<=[۔.;؛])\s+', text)
    segments = [s.strip() for s in raw if s.strip()]
    out = []
    for s in segments:
        if len(s) <= max_chars:
            out.append(s)
        else:
            # Split long segment by comma/clause (، ؛) then by length
            sub = re.split(r'(?<=[،؛,])\s+', s)
            cur = []
            cur_len = 0
            for part in sub:
                part = part.strip()
                if not part:
                    continue
                if cur_len + len(part) + 2 <= max_chars and cur:
                    cur.append(part)
                    cur_len += len(part) + 2
                else:
                    if cur:
                        out.append('، '.join(cur))
                    cur = [part]
                    cur_len = len(part)
            if cur:
                out.append('، '.join(cur))
    return out if out else [text]


def build_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    except ImportError:
        print('Install: pip install reportlab')
        raise

    # Register Vazir so Persian (and Latin) render correctly instead of rectangles
    vazir_path = os.path.join(PROJECT_ROOT, 'fonts', 'Vazir.ttf')
    if os.path.isfile(vazir_path):
        pdfmetrics.registerFont(TTFont('Vazir', vazir_path))
        pdf_font = 'Vazir'
    else:
        pdf_font = 'Helvetica'  # fallback if font missing

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    doc = SimpleDocTemplate(PDF_PATH, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    # RTL paragraph: wordWrap='RTL' + alignment=TA_RIGHT so Persian lays out right-to-left
    title_style = ParagraphStyle(name='Title', parent=styles['Heading1'], fontName=pdf_font, fontSize=16, spaceAfter=12, alignment=TA_RIGHT, wordWrap='RTL')
    heading_style = ParagraphStyle(name='Heading2', parent=styles['Heading2'], fontName=pdf_font, fontSize=12, spaceAfter=8, alignment=TA_RIGHT, wordWrap='RTL')
    body_style = ParagraphStyle(name='Body', parent=styles['Normal'], fontName=pdf_font, fontSize=9, spaceAfter=4, alignment=TA_RIGHT, wordWrap='RTL')
    small_style = ParagraphStyle(name='Small', parent=styles['Normal'], fontName=pdf_font, fontSize=8, spaceAfter=2, alignment=TA_RIGHT, wordWrap='RTL')

    sl = SECTION_LABELS
    story = []
    story.append(Paragraph(_rtl_reshape_paragraph(f"Training Level Information – Export for Confirmation | {sl['training_level_info'][1]}"), title_style))
    story.append(Spacer(1, 0.2 * inch))

    # --- Section 1: Training Level Info (EN + FA labels), RTL + reshape for Persian ---
    story.append(Paragraph(_rtl_reshape_paragraph(f"1. {sl['training_level_info'][0]} | {sl['training_level_info'][1]}"), heading_style))
    story.append(Spacer(1, 0.1 * inch))

    for level_key, level_data in TRAINING_LEVELS_DATA.items():
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>Level | سطح: {level_data['label_en']} | {level_data['label_fa']}</b>"), body_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['description'][0]} | {sl['description'][1]}</b>"), small_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"EN: {level_data['description_en']}"), body_style))
        # FA description as separate short RTL paragraphs for clean layout
        for seg in _split_persian_description(level_data['description_fa']):
            story.append(Paragraph(_rtl_reshape_paragraph(seg), body_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['goals'][0]} | {sl['goals'][1]}</b>"), small_style))
        for g in level_data['goals']:
            story.append(Paragraph(_rtl_reshape_paragraph(f"• {g['en']} | {g['fa']}"), body_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['features_per_purpose'][0]} | {sl['features_per_purpose'][1]}</b>"), small_style))
        for purpose_key, purpose_label in PURPOSE_LABELS.items():
            p = level_data['purposes'][purpose_key]
            story.append(Paragraph(_rtl_reshape_paragraph(f"  {purpose_label[0]} | {purpose_label[1]}:"), small_style))
            story.append(Paragraph(_rtl_reshape_paragraph(f"    Sessions/week: {p.get('sessions_per_week', '')} | Sets: {p.get('sets_per_action', '')} | Reps: {p.get('reps_per_action', '')} | Break: {p.get('break_between_sets', '')}"), small_style))
            story.append(Paragraph(_rtl_reshape_paragraph(f"    Focus EN: {p.get('training_focus_en', '')}"), small_style))
            story.append(Paragraph(_rtl_reshape_paragraph(f"    Focus FA: {p.get('training_focus_fa', '')}"), small_style))
        story.append(Spacer(1, 0.15 * inch))

    story.append(PageBreak())
    story.append(Paragraph(_rtl_reshape_paragraph(f"2. {sl['corrective_movements'][0]} | {sl['corrective_movements'][1]}"), heading_style))
    story.append(Spacer(1, 0.1 * inch))

    for injury_key, data in INJURIES_DATA.items():
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>Injury | آسیب: {data['label_en']} | {data['label_fa']}</b>"), body_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['purposes'][0]} | {sl['purposes'][1]}</b>"), small_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"EN: {data['purposes_en']}"), body_style))
        for seg in _split_persian_description(data['purposes_fa']):
            story.append(Paragraph(_rtl_reshape_paragraph(seg), body_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['allowed_movements'][0]} | {sl['allowed_movements'][1]}</b>"), small_style))
        for m in data['allowed_movements']:
            story.append(Paragraph(_rtl_reshape_paragraph(f"• {m['en']} | {m['fa']}"), body_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['forbidden_movements'][0]} | {sl['forbidden_movements'][1]}</b>"), small_style))
        for m in data['forbidden_movements']:
            story.append(Paragraph(_rtl_reshape_paragraph(f"• {m['en']} | {m['fa']}"), body_style))
        story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['important_notes'][0]} | {sl['important_notes'][1]}</b>"), small_style))
        if data.get('important_notes_en'):
            story.append(Paragraph(_rtl_reshape_paragraph(f"EN: {data['important_notes_en']}"), body_style))
        for seg in _split_persian_description(data.get('important_notes_fa', '')):
            if seg.strip():
                story.append(Paragraph(_rtl_reshape_paragraph(seg), body_style))
        story.append(Spacer(1, 0.12 * inch))

    story.append(Paragraph(_rtl_reshape_paragraph(f"<b>{sl['common_note'][0]} | {sl['common_note'][1]}</b>"), heading_style))
    story.append(Paragraph(_rtl_reshape_paragraph(f"EN: {COMMON_INJURY_NOTE['en']}"), body_style))
    for seg in _split_persian_description(COMMON_INJURY_NOTE['fa']):
        story.append(Paragraph(_rtl_reshape_paragraph(seg), body_style))

    doc.build(story)
    print('PDF saved:', PDF_PATH)


def main():
    print('Generating Excel and PDF exports...')
    build_excel()
    build_pdf()
    print('Done. Please review:')
    print('  -', EXCEL_PATH)
    print('  -', PDF_PATH)


if __name__ == '__main__':
    main()
